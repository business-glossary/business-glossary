#!/usr/bin/python
# -*- encoding: utf-8-*-

import os
import logging
import csv

from openpyxl import load_workbook

from app.core import create_app
from app.config import config, BASE_DIR
from app.models import db
from app.models import Location, Table, Column

LOGGER = logging.getLogger("business-glossary.load_data")


def load_column_desc_excel(file_name):
    wb = load_workbook(file_name)

    sheet = wb.get_sheet_by_name('Tables')

    table_data = []

    print("reading rows")

    if sheet['A1'].value.upper() != 'LIBRARY' or \
       sheet['B1'].value.upper() != 'TABLE' or \
    sheet['D1'].value.upper() != 'LABEL':
        print("Expecting a worksheet named Tables with columns library, table and label.")
        quit()

    for row in range(2, sheet.max_row + 1):
        library = sheet['A' + str(row)].value
        table = sheet['B' + str(row)].value
        label = sheet['D' + str(row)].value

        row_data = {
            'library': library,
            'table': table,
            'description': label
        }

        table_data.append(row_data)

    return table_data


def load_column_desc_csv(file_name):
    LOGGER.info("Updating table description")

    with open(file_name, 'rt') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            LOGGER.info("Updating description for table %s.%s", row['library'], row['table'])

            t = Table.query.join(Location).filter(Location.name == row['library'], Table.name == row['table']).first()
            if t:
                t.description = row['description']
                db.session.commit()
            if not t:
                logger.info("Could not find the table %s.%s", row['library'], row['table'])             


def update_table_description(table_dict):
    LOGGER.info("Updating table description")

    for row in table_dict:
        LOGGER.info("Updating description for table %s.%s", row['library'], row['table'])
        t = Table.query.join(Location).filter(Location.name == row['library'], Table.name == row['table']).first()
        if t:
            t.description = row['description']
            db.session.commit()
            if row['table'] == 'LGD_ACCOUNT':
                print (row['description'])
        if not t:
            LOGGER.info("Could not find the table %s.%s", row['library'], row['table'])


if __name__ == "__main__":
    LOG_FORMAT = "%(asctime)-15s [%(levelname)s] %(message)s"
    logging.basicConfig(level=logging.INFO, format=LOG_FORMAT)

    LOGGER.info("Load process started")

    app = create_app(os.getenv('BG_CONFIG') or 'default')
    
    with app.app_context():

        app.config['SQLALCHEMY_ECHO'] = False
        app.config['SQLALCHEMY_DEBUG'] = False

        # Interface files are placed in a directory name bg_interface at the same level
        # as the application directory, i.e.
        #
        #   - bg_interface
        #   - business_glossary
        #
        # Call os.path.dirname twice to walk up to the parent directory

        FILE_PATH = os.path.join(os.path.dirname(BASE_DIR), 'bg_interface')

        # update_table_description(os.path.join(FILE_PATH, "bg_tables.csv"))

        table_data = load_column_desc_excel('G:\Risk Analytics\Governance\Business Glossary\Data Dictionary\Tables.xlsx')
        # table_data = load_column_desc_csv('G:\Risk Analytics\Governance\Business Glossary\Data Dictionary\Tables.xlsx')

        update_table_description(table_data)

        LOGGER.info("Load process ended")
