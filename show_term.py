import argparse

from app.core import create_app
from app.config import config

application = create_app('production').app_context().push()

from app.main.models import Term, Rule, Note

from openpyxl import Workbook

#########################################################################################
##                                                                                     ##
##  Simply pipe delimited lists                                                        ##
##                                                                                     ##
#########################################################################################

def list_terms():
    for term in Term.query.all():
        print(term.name, end="| ")

def list_rules():
    for rule in Rule.query.all():
        print(rule.name, end="|")
        for term in rule.terms:
            print(term.name, end="|")
            #print("Term: %s" % term.name )
        print()

#########################################################################################
##                                                                                     ##
##  Term, rule, rule note dump with indent                                             ##
##                                                                                     ##
#########################################################################################

def list_terms_rules_indent():
    '''
    Produce a list of terms, rules and rule notes in this format:

    Term: Application Number
      Rule: Balance and Account Balance Correction
        Rule Note: Source: `acct\2.0_account_history\acct_hist_ext.sas`
    '''
    for term in Term.query.all():
#       print(term.name, end="|")
        print("Term: %s" % term.name)
        for rule in term.rules:
#           print(rule.name, end="|")
            print("  Rule: %s" % rule.name )
            for rn in rule.comments:
                print("    Rule Note: %s" % rn.note.split('\n', 1)[0])
       
#########################################################################################
##                                                                                     ##
##  Functions for the term dump to Excel                                               ##
##                                                                                     ##
#########################################################################################

def output_comments(ws, rule, start_row):
    if rule.comments.count() > 0:
        print("    >> There are %s rule notes" % rule.comments.count())
        comment_row = start_row
        for rn in rule.comments:
            ws.cell(row=comment_row, column=3).value = rn.note.split('\n', 1)[0]
            print("    Rule Note: %s (%s)" % (rn.note.split('\n', 1)[0], comment_row))
            comment_row += 1
        return comment_row - 1
    else:
        print("    >> No rule comments, not adding one to row")
        return start_row

def output_rules(ws, term, start_row):
    if term.rules:
        print("  >> There are rules")
        rule_row = start_row
        for rule in term.rules:
            ws.cell(row=rule_row, column=2).value = rule.name
            print("  Rule: %s (%s)" % (rule.name, rule_row))
            # Do rule comments here
            rule_row = output_comments(ws, rule, rule_row)
            # Add one?
            print("  >> Add one to rule row %s" % rule_row)
            rule_row += 1

        # Is the final + 1 needed?
        return rule_row
    else:
        return start_row + 1

def output_terms(ws, start_row):
    term_row = start_row
    for term in Term.query.order_by(Term.name).all():
        ws.cell(row=term_row, column=1).value = term.name
        print("Term: %s (%s)" % (term.name, term_row))
        term_row = output_rules(ws, term, term_row)

def excel_term_dump():
    wb = Workbook()
    ws = wb.active

    output_terms(ws, 1)

    wb.save("H:\glossary.xlsx")

#########################################################################################
##                                                                                     ##
##  Print routine                                                                      ##
##                                                                                     ##
#########################################################################################

def print_report():

    from markdown import markdown
    import pdfkit
    import flask

    output_filename = 'glossary.pdf'

    #terms = Term.query.limit(10)
    terms = Term.query.order_by(Term.name).all()

    html_text = flask.render_template('print/glossary_print.html', terms=terms)

    options = {
        'page-size': 'A4',
        'dpi': 300,
        'margin-top': '20mm',
        'margin-right': '20mm',
        'margin-bottom': '20mm',
        'margin-left': '20mm',
        'encoding': "UTF-8",
        'header-left': "BOQ Credit Risk",
        'header-right': "Full Glossary",
        'header-font-name': 'Roboto',
        'header-font-size': '8',
        'header-spacing': '10',
        'footer-left': '[date], [time]',
        'footer-right': 'Page [page] of [topage]',
        'footer-font-name': 'Roboto',
        'footer-font-size': '8',
        'footer-spacing': '10',
        'no-outline': None
    }

    css = 'C:/Users/Alan/Projects/glossary/print-test/style.css'

    print(html_text)

    pdfkit.from_string(html_text, output_filename, options=options, css=css)


#########################################################################################
##                                                                                     ##
##  Argparse custom usage function                                                     ##
##                                                                                     ##
#########################################################################################

    
def msg(name=None):                                                            
    return '''show_term.py command
    
        Commands are:

        rules              Pipe delimited dump of rules
        term               Pipe delimited dump of terms
        term_rules_indent  Dump terms, rules and rule notes in indented form
        excel_term_dump    Dump terms, rules and rule notes to Excel
        '''
   
    
#########################################################################################
##                                                                                     ##
##  Main                                                                               ##
##                                                                                     ##
#########################################################################################

if __name__ == "__main__":

    parser = argparse.ArgumentParser(description='Utility to dump data from the Glossary',
                                     usage=msg())

    parser.add_argument("command", help="Enter the command to run")
    args = parser.parse_args()

    if args.command == "rules":
        list_rules()
    elif args.command == "terms":
        list_terms()
    elif args.command == "excel_term_dump":
        excel_term_dump()
    elif args.command == "terms_rules_indent":
        list_terms_rules_indent()
    elif args.command == "print":
        print_report()
    else:
        parser.print_help()
